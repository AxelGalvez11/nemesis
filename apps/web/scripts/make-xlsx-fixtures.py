"""
Build the XLSX fixture set — the failure surface, before any parser exists.

🔴 FIXTURES FIRST, AND THE MODEL AFTERWARDS (owner, 2026-08-12). The point is not
to have files to test against; it is to find out what a real workbook actually
contains before deciding what a spreadsheet cell IS. Every field the parser ends
up carrying should be traceable to something one of these files does.

REAL FILES, GENERATED. openpyxl writes genuine .xlsx archives, and where the case
needs a program that actually computes — cached values, an "ugly export" — the
file is round-tripped through LibreOffice, which is a real spreadsheet
application doing its real job. Nothing here is a mock of a zip.

  uv run --with openpyxl python scripts/make-xlsx-fixtures.py <outdir>
"""

import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

out = Path(sys.argv[1] if len(sys.argv) > 1 else "lib/notebooks/fixtures/xlsx")
out.mkdir(parents=True, exist_ok=True)


def save(wb, name):
    path = out / name
    wb.save(path)
    print(f"  wrote {name}")
    return path


# ── A. the ordinary workbook, with the ordinary nastiness ──────────────────
# Multiple sheets, a DUPLICATE header, blank cells, and a sheet that is empty.
wb = Workbook()
grades = wb.active
grades.title = "Grades"
# 🔴 TWO COLUMNS CALLED "Quiz 1". Real exports do this constantly, and a parser
# that keys columns by name silently loses one of them.
grades.append(["Student", "Quiz 1", "Quiz 1", "Final"])
grades.append(["Ada", 88, 91, 90])
grades.append(["Grace", None, 78, 80])   # a genuinely blank cell mid-row
grades.append(["Alan", 95, None, 97])
notes = wb.create_sheet("Notes")
notes.append(["Marking scheme"])
notes.append(["Quizzes are averaged; the final is weighted 60%."])
wb.create_sheet("Empty Sheet")           # no cells at all
# 🔴 A HIDDEN SHEET IS A HIDDEN UNIT, not a hidden row — and a new DocUnit field
# dies at the validator's field-by-field rebuild unless it is added there too.
# This fixture exists to make that failure reproducible rather than theoretical.
internal = wb.create_sheet("Internal")
internal.append(["Moderation notes", "do not show students"])
internal.sheet_state = "hidden"
save(wb, "basic.xlsx")

# ── H. number formats: what "displayed value" would even mean ───────────────
# 🔴 MEASURED, NOT ASSUMED. A date is stored as a serial number and a percentage
# as a fraction; only the format string turns them into what a human sees.
# Applying those formats is a rendering engine we are not building in this slice,
# so this fixture exists to record exactly how wide the gap is.
wb = Workbook()
ws = wb.active
ws.title = "Formats"
ws.append(["What", "Stored as"])
from datetime import date
ws["A2"] = "a date"
ws["B2"] = date(2026, 3, 15)
ws["A3"] = "a percentage"
ws["B3"] = 0.075
ws["B3"].number_format = "0.0%"
ws["A4"] = "money"
ws["B4"] = 1234.5
ws["B4"].number_format = '"$"#,##0.00'
ws["A5"] = "plain text that looks numeric"
ws["B5"] = "007"
ws["A6"] = "grouped number"
ws["B6"] = 1234567.891
ws["B6"].number_format = "#,##0.00"
ws["A7"] = "builtin percent"
ws["B7"] = 0.5
ws["B7"].number_format = "0%"          # builtin numFmtId 9
ws["A8"] = "a format we will not render"
ws["B8"] = 12
ws["B8"].number_format = '0" widgets"'  # deliberately outside the supported classes
save(wb, "number-formats.xlsx")

# ── H2. a FORMULA whose cached result carries a number format ──────────────
# 🔴 THE CASE THAT JOINS THE TWO HALVES. A percentage is usually computed, not
# typed, so the display rule has to apply to a formula's cached result too —
# otherwise every derived percentage in every workbook shows as a fraction.
# Written by openpyxl (formula, no cached value), then calculated by LibreOffice.
wb = Workbook()
ws = wb.active
ws.title = "Derived"
ws.append(["Cohort", "Sat", "Passed", "Pass rate", "Fees"])
ws.append(["2026", 40, 31, "=C2/B2", "=B2*125"])
ws["D2"].number_format = "0.0%"
ws["E2"].number_format = '"$"#,##0.00'
save(wb, "formatted-formulas.xlsx")

# ── B. formulas with NO cached value ───────────────────────────────────────
# openpyxl does not calculate. This is exactly what a workbook written by a
# program looks like: <f> present, <v> absent. If the parser reads only values,
# this file is silently blank.
wb = Workbook()
ws = wb.active
ws.title = "Totals"
ws.append(["Item", "Qty", "Unit", "Cost"])
ws.append(["Bolts", 10, 0.25, "=B2*C2"])
ws.append(["Nuts", 40, 0.10, "=B3*C3"])
ws.append(["Total", None, None, "=SUM(D2:D3)"])
ws["A6"] = "=CONCATENATE(A2,\" and \",A3)"
save(wb, "formulas-uncalculated.xlsx")

# ── D. merged headers, hidden row, hidden column ───────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Schedule"
# A merged banner across four columns, then a merged two-column group header.
ws["A1"] = "Autumn Term"
ws.merge_cells("A1:D1")
ws["A2"] = "Week"
ws["B2"] = "Assessment"
ws.merge_cells("B2:C2")
ws["D2"] = "Notes"
ws.append(["1", "Quiz", "10%", "in class"])
ws.append(["2", "Essay", "30%", "submit online"])
ws.append(["3", "Lab", "20%", "hidden row below"])
ws.append(["4", "SECRET", "0%", "this row is hidden"])
ws.row_dimensions[6].hidden = True        # the 4/SECRET row
ws.column_dimensions["C"].hidden = True   # the weighting column
save(wb, "merged-hidden.xlsx")

# ── E. 🔴 THE USED RANGE DOES NOT START AT A1 ──────────────────────────────
# THE ONE DEFECT THIS SLICE COULD SHIP SILENTLY. A grid coordinate is not a cell
# reference unless you know where the grid begins. A round-trip test comparing
# stored cells to read-back cells cannot see the error, because both sides are
# wrong in the same way. The assertion has to be the A1 STRING.
wb = Workbook()
ws = wb.active
ws.title = "Offset"
ws["C5"] = "Drug"
ws["D5"] = "Dose"
ws["C6"] = "Amoxicillin"
ws["D6"] = "500 mg"
ws["C7"] = "Ibuprofen"
ws["D7"] = "400 mg"
save(wb, "offset-origin.xlsx")

# ── F. a real Excel table (ListObject) over a range ────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Inventory"
ws.append(["Code", "Name", "Count"])
ws.append(["A1", "Beaker", 12])
ws.append(["A2", "Flask", 7])
ws.append(["A3", "Pipette", 30])
table = Table(displayName="Inventory", ref="A1:C4")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws.add_table(table)
save(wb, "list-object.xlsx")

# ── J. 🔴 OFFSET *AND* HIDDEN TOGETHER ─────────────────────────────────────
# The origin arithmetic applies to hidden row/column indexes too, and neither
# earlier fixture exercises that: `merged-hidden` starts at A1 (so subtracting
# the origin is a no-op) and `offset-origin` hides nothing. A wrong subtraction
# here lands negative and is silently filtered away — a hidden row that simply
# disappears, which is the same shape as the A1/C5 defect one field over.
wb = Workbook()
ws = wb.active
ws.title = "Offset Hidden"
ws["C5"] = "Drug"
ws["D5"] = "Dose"
ws["E5"] = "Note"
ws["C6"] = "Amoxicillin"
ws["D6"] = "500 mg"
ws["E6"] = "with food"
ws["C7"] = "WITHDRAWN"
ws["D7"] = "-"
ws["E7"] = "this row is hidden"
ws["C8"] = "Ibuprofen"
ws["D8"] = "400 mg"
ws["E8"] = "as needed"
ws.row_dimensions[7].hidden = True     # sheet row 7 == grid row 2
ws.column_dimensions["D"].hidden = True  # sheet column D == grid column 1
save(wb, "offset-hidden.xlsx")

# ── K. two sheets, each with its OWN table, at the SAME range ──────────────
# 🔴 THE DISCRIMINATING CASE FOR TABLE OWNERSHIP. Both tables start at A1, so a
# reader that falls back to "every table part in the archive" cannot tell them
# apart and will attach one sheet's header count and name to the other.
wb = Workbook()
first = wb.active
first.title = "Reagents"
first.append(["Code", "Name"])
first.append(["R1", "Ethanol"])
first.append(["R2", "Acetone"])
first.add_table(Table(displayName="Reagents", ref="A1:B3"))
second = wb.create_sheet("Glassware")
second.append(["Code", "Name"])
second.append(["G1", "Beaker"])
second.append(["G2", "Funnel"])
second.add_table(Table(displayName="Glassware", ref="A1:B3"))
save(wb, "two-tables.xlsx")

# ── I. a chart: content this reader does NOT turn into text ────────────────
# 🔴 UNSUPPORTED IS NOT ABSENT (owner, 2026-08-12). Charts are deliberately out
# of scope, and the wrong way to be out of scope is to say nothing — a workbook
# whose point is its chart would report itself completely read. This fixture
# exists so that claim is testable rather than asserted.
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws.title = "Results"
ws.append(["Cohort", "Passed"])
for row in [["2024", 41], ["2025", 55], ["2026", 62]]:
    ws.append(row)
chart = BarChart()
chart.title = "Passes by cohort"
chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
ws.add_chart(chart, "D2")
save(wb, "with-chart.xlsx")

# ── C + G. what a real spreadsheet program produces ────────────────────────
soffice = shutil.which("soffice") or shutil.which("libreoffice")
if not soffice:
    print("  ⚠ LibreOffice not found — skipping the calculated and stale fixtures")
    sys.exit(0)

def calculate(source_name, target_name):
    """Round-trip through LibreOffice so a real program computes the formulas."""
    work = out / "_convert"
    work.mkdir(exist_ok=True)
    subprocess.run(
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(work), str(out / source_name)],
        check=True, capture_output=True, timeout=180,
    )
    target = out / target_name
    shutil.move(str(work / source_name), target)
    shutil.rmtree(work)
    print(f"  wrote {target_name} (LibreOffice computed the values)")
    return target


target = calculate("formulas-uncalculated.xlsx", "formulas-calculated.xlsx")
# The formatted-formula fixture is only interesting once something has computed
# it: an uncalculated percentage has no cached result to display.
calculate("formatted-formulas.xlsx", "formatted-formulas.xlsx")

# G. 🔴 A CACHED VALUE THAT DISAGREES WITH ITS FORMULA.
# Edit an input cell in the raw XML and do NOT recalculate — which is precisely
# what any program that writes a spreadsheet without a formula engine does. The
# stored <v> for the total is now stale. The owner's rule: preserve BOTH, never
# silently pick one as truth.
stale = out / "formulas-stale.xlsx"
with zipfile.ZipFile(target) as zin:
    names = zin.namelist()
    blobs = {n: zin.read(n) for n in names}
sheet = next(n for n in names if n.startswith("xl/worksheets/sheet"))
xml = blobs[sheet].decode("utf-8")
# Qty of bolts: 10 → 9999. Every dependent cached total is now wrong.
before = xml
xml = xml.replace("<v>10</v>", "<v>9999</v>", 1)
if xml == before:
    print("  ⚠ could not find the input value to make stale — skipping")
else:
    with zipfile.ZipFile(stale, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, xml.encode("utf-8") if n == sheet else blobs[n])
    print("  wrote formulas-stale.xlsx (cached totals disagree with the formulas)")

print(f"\n{len(list(out.glob('*.xlsx')))} fixtures in {out}")
