"""
Build the CSV fixture set — the failure surface, before any parser exists.

🔴 FIXTURES FIRST, AND THE MODEL AFTERWARDS. The XLSX slice proved this twice:
the two defects that reached production were both found by a file a SECOND
program wrote (`hidden="true"` where our own library wrote `hidden="1"`; an
escaped `\\$` where it wrote `"$"`). Our code and our own generated fixtures
agreed with each other, which is exactly the state in which a test suite tells
you nothing.

So this writes CSVs three ways:

  1. Python's stdlib `csv` — a real RFC 4180 writer (CRLF, minimal quoting)
  2. LibreOffice, converting a real .xlsx — an independent implementation with
     its own opinions about quoting, line endings and locale
  3. RAW BYTES, for the things no library will emit on request: a UTF-8 BOM, a
     lone-LF file, ragged rows, a semicolon export, a stray CR.

  uv run --with openpyxl python scripts/make-csv-fixtures.py <outdir>
"""

import csv
import io
import shutil
import subprocess
import sys
from pathlib import Path

out = Path(sys.argv[1] if len(sys.argv) > 1 else "lib/notebooks/fixtures/csv")
out.mkdir(parents=True, exist_ok=True)


def write_raw(name: str, data: bytes):
    (out / name).write_bytes(data)
    print(f"  wrote {name}  ({len(data)} bytes, raw)")


def write_rows(name: str, rows, **kwargs):
    """Python's own csv writer. newline='' so it controls the line endings."""
    buffer = io.StringIO(newline="")
    csv.writer(buffer, **kwargs).writerows(rows)
    data = buffer.getvalue().encode("utf-8")
    (out / name).write_bytes(data)
    print(f"  wrote {name}  ({len(data)} bytes, python csv)")


# ── A. the ordinary export, with the ordinary nastiness ────────────────────
# 🔴 EVERY ONE OF THESE IS A DIFFERENT WAY TO LOSE A CELL. A comma inside quotes
# is not a delimiter; a doubled quote is one character; a newline inside quotes
# does not end the row. A parser that splits on commas and newlines gets all
# three wrong and produces a grid that looks entirely plausible.
write_rows(
    "quoting.csv",
    [
        ["Student", "Note", "Score"],
        ["Ada", "Comma, inside a field", "88"],
        ["Grace", 'She said ""well done""'.replace('""', '"'), "91"],
        ["Alan", "Two\nlines in one cell", "95"],
        ["Edsger", "", "70"],            # a genuinely empty field
        ["Barbara", "trailing comma next", ""],
    ],
)

# ── B. duplicate headers and a blank first column ──────────────────────────
# Real exports do this constantly. A parser that keys by column NAME silently
# loses one of the two "Quiz 1" columns.
write_rows(
    "duplicate-headers.csv",
    [
        ["", "Quiz 1", "Quiz 1", "Final"],
        ["Ada", "88", "91", "90"],
        ["Grace", "", "78", "80"],
    ],
)

# ── C. 🔴 RAGGED ROWS — the file disagrees with itself about its own width ──
# Nothing in the format forbids it, and real files do it: a trailing summary
# line, a note appended under a table, a row whose last empty fields were
# trimmed by whatever wrote it. The grid must be as wide as the WIDEST row, and
# the short rows must be short — padding them with "" invents cells, and
# truncating the long one deletes data.
write_raw(
    "ragged.csv",
    b"Code,Name,Count\r\n"
    b"B1,Beaker,12\r\n"
    b"B2,Burette\r\n"                     # one field short
    b"B3,Flask,7,extra,fields\r\n"        # two fields long
    b"Totals\r\n",                        # a lone trailing cell
)

# ── D. 🔴 A UTF-8 BOM. This is what Excel writes, and it is invisible. ──────
# Left in place, the first header becomes "﻿Student" — which reads
# identically in every log and error message, and never matches "Student".
write_raw(
    "bom-crlf.csv",
    "﻿".encode("utf-8") + b"Student,Grade\r\nAda,A\r\nGrace,B\r\n",
)

# ── E. lone LF, no trailing newline, non-ASCII ─────────────────────────────
write_raw(
    "lf-unicode.csv",
    "Term,Definition\nrésumé,a summary\nnaïve,untrained\n"
    "日本語,the Japanese language".encode("utf-8"),
)

# ── F. 🔴 A SEMICOLON EXPORT. Not a guess to make — a fact to preserve. ────
# Excel writes semicolons wherever the decimal separator is a comma. The point
# of this fixture is NOT to justify sniffing: it is to make sure that whatever
# we decide, the decision is visible and testable rather than accidental.
write_raw(
    "semicolon.csv",
    b"Datum;Thema;Gewichtung\r\n"
    b"2026-03-15;Einf\xc3\xbchrung;10%\r\n"
    b"2026-03-22;Vertiefung;20%\r\n",
)

# ── F2. 🔴 THE FIXTURES THAT CALIBRATE THE DELIMITER RULE ──────────────────
# A comma file and a semicolon file are both UNAMBIGUOUS — for each, exactly one
# candidate yields a consistent grid, so neither can tell whether our rule is
# evidence or a coin flip. These two do.
#
# `ambiguous-comma-wins`: BOTH comma and semicolon split it into a consistent
# 2-column grid, and they disagree about every cell. This is not exotic — it is
# what any ordinary comma CSV looks like when each row happens to contain one
# semicolon. Comma must win here BY STATED PRIORITY, not by luck of iteration
# order: it is the format's namesake, and a rule that refused this would refuse
# real files.
write_raw(
    "ambiguous-comma-wins.csv",
    b"Topic,Note\r\nlists;arrays,see chapter 2\r\nqueues;stacks,see chapter 3\r\n",
)

# `ambiguous-no-comma`: comma is INCONSISTENT (1 field, then 3, then 1), and TWO
# other candidates — semicolon and tab — each yield a consistent 3-column grid
# that contradicts the other. There is no evidence here, only a choice, so the
# reader must decline to split rather than pick one. The rows survive verbatim;
# what is refused is the claim to know where the columns are.
write_raw(
    "ambiguous-no-comma.csv",
    b"a;b\tc\r\nd;e\tf,g\r\nh;i\tj\r\n",
)

# ── G. a quoted field containing CRLF, and quotes at the very edges ────────
write_raw(
    "quoted-crlf.csv",
    b'Item,Description\r\n'
    b'A,"first line\r\nsecond line"\r\n'
    b'B,"""fully quoted"""\r\n'
    b'C,"ends with a quote """\r\n',
)

# ── H. blank lines, and a file that is only blank lines ────────────────────
# 🔴 A BLANK LINE IS A ROW, not a separator, when it sits between data rows —
# deleting it renumbers every row beneath it, and a citation to row 5 then
# points at row 4.
write_raw("blank-lines.csv", b"a,b\r\n\r\nc,d\r\n\r\n\r\ne,f\r\n")
write_raw("empty.csv", b"")
write_raw("newlines-only.csv", b"\r\n\r\n")

# ── I. a single cell, and a single column ──────────────────────────────────
write_raw("one-cell.csv", b"just one value\r\n")

# ── J. leading/trailing spaces, and a numeric-looking string ───────────────
# 🔴 " 007 " IS THE AUTHOR'S TEXT. Trimming it, or reading it as 7, are both
# this parser deciding it knows better than the file.
write_raw(
    "spaces-and-numbers.csv",
    b'Code, Padded ,Value\r\n007, spaced ,0.075\r\n"  quoted spaces  ",x,1e3\r\n',
)

# ── K. what a real spreadsheet program produces ────────────────────────────
# 🔴 THE INDEPENDENT WRITER. LibreOffice quotes differently from Python, picks
# its own line endings, and re-renders values through its own display rules.
# This is the fixture most likely to disagree with our assumptions.
soffice = shutil.which("soffice") or shutil.which("libreoffice")
if not soffice:
    print("  ⚠ LibreOffice not found — skipping the independently-written fixture")
    sys.exit(0)

try:
    from openpyxl import Workbook
except ImportError:
    print("  ⚠ openpyxl not available — skipping the independently-written fixture")
    sys.exit(0)

wb = Workbook()
ws = wb.active
ws.title = "Export"
ws.append(["Cohort", "Note", "Rate", "Fees"])
ws.append(["2026", "Comma, and \"quotes\"", 0.775, 5000])
ws.append(["2025", "Two\nlines", 0.075, 4500])
ws["C2"].number_format = "0.0%"
ws["C3"].number_format = "0.0%"
ws["D2"].number_format = '"$"#,##0.00'
ws["D3"].number_format = '"$"#,##0.00'
source = out / "_soffice-source.xlsx"
wb.save(source)

work = out / "_convert"
work.mkdir(exist_ok=True)
subprocess.run(
    [soffice, "--headless", "--convert-to", "csv", "--outdir", str(work), str(source)],
    check=True, capture_output=True, timeout=180,
)
shutil.move(str(work / "_soffice-source.csv"), out / "libreoffice-export.csv")
shutil.rmtree(work)
source.unlink()
data = (out / "libreoffice-export.csv").read_bytes()
print(f"  wrote libreoffice-export.csv  ({len(data)} bytes, LibreOffice)")

print(f"\n{len(list(out.glob('*.csv')))} fixtures in {out}")
